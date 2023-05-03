
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 12
ws['B1'] = 12
ws['C1'] = "=SUM(A1:B1)"
ws['D1'] = "=A1 * B1"
wb.save("sample.xls")

import openpyxl

from openpyxl.chart import BarChart, Reference
wb = openpyxl.Workbook()
sheet = wb.active

for i in range(10):
    sheet.append([i])

values = Reference(sheet, min_col = 1, min_row = 1, max_col = 1, max_row = 10)
chart = BarChart()


chart.add_data(values)
chart.title = "MY Bar Chart"
chart.x_axis.title = "X-Axis"
chart.y_axis.title = "Y-Axis"

sheet.add_chart(chart, "E2")

wb.save("barchart.xlsx")

